"""
main.py
=======
Entry point chính của chương trình.
  • Khởi tạo GPTExtractor
  • Chạy CurriculumExtractor (agent.py)
  • Xuất báo cáo Excel + PDF (ExcelReporter / PDFReporter)
  • Import vào Neo4j (database.py)

Cài đặt:
    pip install pdfplumber neo4j openpyxl reportlab openai

Cấu hình qua biến môi trường:
    export PDF_FILE_PATH="./chuong_trinh_dao_tao.pdf"
    export NEO4J_URL="neo4j+s://xxxx.databases.neo4j.io"
    export NEO4J_USER="xxxx"
    export NEO4J_PASSWORD="your_password"
    export GEMINI_API_KEY="AIza..."  # tùy chọn, lấy miễn phí tại aistudio.google.com
"""

import os

from dotenv import load_dotenv
load_dotenv()  # đọc file .env cùng thư mục

from extractor import GPTExtractor, CurriculumExtractor
from database import Neo4jImporter

# ── Export libs ───────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, PageBreak, HRFlowable)
from reportlab.lib.enums import TA_CENTER

# ================= CẤU HÌNH =================
PDF_FILE_PATH  = os.environ.get('PDF_FILE_PATH',  './chuong_trinh_dao_tao.pdf')
NEO4J_URL      = os.environ.get('NEO4J_URL',      'neo4j+s://a254bc54.databases.neo4j.io')
NEO4J_USER     = os.environ.get('NEO4J_USER',     'a254bc54')
NEO4J_PASSWORD = os.environ.get('NEO4J_PASSWORD', '')
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
# ============================================


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  EXCEL REPORTER                                                             ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class ExcelReporter:
    """Xuất báo cáo curriculum ra file .xlsx (6 sheet)."""

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL    = PatternFill("solid", fgColor="DEEAF1")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    TITLE_FONT  = Font(bold=True, color="1F4E79", name="Arial", size=13)
    NORMAL_FONT = Font(name="Arial", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    _t          = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=_t, right=_t, top=_t, bottom=_t)

    def _hdr(self, ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = self.HEADER_FONT; c.fill = self.HEADER_FILL
        c.alignment = self.CENTER; c.border = self.BORDER
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width

    def _cell(self, ws, row, col, value, alt=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = self.NORMAL_FONT; c.alignment = self.LEFT; c.border = self.BORDER
        if alt:
            c.fill = self.ALT_FILL
        return c

    # ── sheets ────────────────────────────────────────────────────────────────

    def _sheet_overview(self, wb, data):
        ws = wb.create_sheet("Tổng quan")
        info = data['program_info']
        ws.merge_cells("A1:C1")
        t = ws["A1"]
        t.value = "THÔNG TIN CHƯƠNG TRÌNH ĐÀO TẠO"
        t.font = self.TITLE_FONT; t.alignment = self.CENTER

        for i, (k, v) in enumerate([
            ("Tên chương trình",  info.get('program_name', '')),
            ("Mã ngành",          info.get('program_code', '')),
            ("Trình độ đào tạo",  info.get('degree_level', '')),
            ("Hình thức đào tạo", info.get('training_form', '')),
            ("Tổng số tín chỉ",   info.get('total_credits', '')),
            ("Thời gian đào tạo", f"{info.get('duration_years', '')} năm"),
            ("Số học kỳ",         info.get('num_semesters', '')),
            ("Năm ban hành",      info.get('year', '')),
        ], start=3):
            alt = i % 2 == 0
            lbl = self._cell(ws, i, 1, k, alt)
            lbl.font = Font(bold=True, name="Arial", size=10)
            self._cell(ws, i, 2, v, alt)

        ws["A12"] = "THỐNG KÊ"; ws["A12"].font = self.TITLE_FONT
        for i, (k, v) in enumerate([
            ("PLO",                   len(data['learning_outcomes'])),
            ("PI",                    len(data['program_indicators'])),
            ("Môn học",               len(data['courses'])),
            ("Học kỳ",                len(data['semesters'])),
            ("Quan hệ tiên quyết",    len(data['prerequisites'])),
            ("Vị trí việc làm",       len(data['job_positions'])),
        ], start=13):
            self._cell(ws, i, 1, k)
            c = self._cell(ws, i, 2, v)
            c.alignment = Alignment(horizontal="center")

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    def _sheet_courses(self, wb, data):
        ws = wb.create_sheet("Danh sách môn học")
        hdrs   = ["Mã HP","Tên học phần","Tín chỉ","Loại","Khoa/Bộ môn",
                  "Tiên quyết","Học trước","Hỗ trợ"]
        widths = [10, 40, 8, 18, 28, 12, 12, 12]
        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            self._hdr(ws, 1, ci, h, w)
        ws.row_dimensions[1].height = 30

        for ri, (cid, c) in enumerate(data['courses'].items(), start=2):
            alt = ri % 2 == 0
            for ci, v in enumerate([
                cid, c['name'], c['credits'],
                "Tự chọn" if c['is_elective'] else "Bắt buộc",
                c['department'], c['prerequisite'],
                c['co_requisite'], c['support_course'],
            ], 1):
                cell = self._cell(ws, ri, ci, v, alt)
                if ci == 3:
                    cell.alignment = Alignment(horizontal="center")

        last = len(data['courses']) + 2
        ws.cell(last, 1, "TỔNG").font = Font(bold=True, name="Arial")
        tc = ws.cell(last, 3, f"=SUM(C2:C{last-1})")
        tc.font = Font(bold=True, name="Arial")
        tc.fill = PatternFill("solid", fgColor="FFD966")

    def _sheet_plo(self, wb, data):
        ws = wb.create_sheet("Chuẩn đầu ra PLO")
        for ci, (h, w) in enumerate(zip(
                ["Mã PLO","Mô tả","Loại","Số PI"], [10, 60, 22, 8]), 1):
            self._hdr(ws, 1, ci, h, w)

        pi_count = {}
        for p in data['program_indicators']:
            pi_count[p['plo_code']] = pi_count.get(p['plo_code'], 0) + 1

        for ri, plo in enumerate(data['learning_outcomes'], 2):
            alt = ri % 2 == 0
            for ci, v in enumerate([plo['code'], plo['description'],
                                     plo['type'], pi_count.get(plo['code'], 0)], 1):
                self._cell(ws, ri, ci, v, alt)

    def _sheet_semesters(self, wb, data):
        ws = wb.create_sheet("Kế hoạch học tập")
        ws.merge_cells("A1:E1")
        t = ws["A1"]
        t.value = "KẾ HOẠCH HỌC TẬP THEO HỌC KỲ"
        t.font = self.TITLE_FONT; t.alignment = self.CENTER
        row = 3

        for sem in sorted(data['semesters'].values(), key=lambda s: s['number']):
            ws.merge_cells(f"A{row}:E{row}")
            hc = ws[f"A{row}"]
            hc.value = f"{sem['name']}  —  {sem['total_credits']} TC  |  Năm {sem['year']}"
            hc.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            hc.fill  = self.SUBHDR_FILL; hc.alignment = self.CENTER
            row += 1

            for ci, (h, w) in enumerate(zip(
                    ["STT","Mã HP","Tên học phần","Tín chỉ","Loại"],
                    [5, 12, 45, 10, 15]), 1):
                self._hdr(ws, row, ci, h, w)
            row += 1

            for stt, cid in enumerate(sem['courses'], 1):
                c   = data['courses'].get(cid, {})
                alt = stt % 2 == 0
                for ci, v in enumerate([
                    stt, cid, c.get('name', ''), c.get('credits', ''),
                    "Tự chọn" if c.get('is_elective') else "Bắt buộc",
                ], 1):
                    self._cell(ws, row, ci, v, alt)
                row += 1
            row += 1

    def _sheet_prerequisites(self, wb, data):
        ws = wb.create_sheet("Quan hệ tiên quyết")
        for ci, (h, w) in enumerate(zip(
                ["Môn học","Tên môn học","Loại quan hệ",
                 "Môn tiên quyết","Tên tiên quyết","Nguồn"],
                [12, 40, 16, 14, 40, 12]), 1):
            self._hdr(ws, 1, ci, h, w)

        for ri, pr in enumerate(data['prerequisites'], 2):
            alt = ri % 2 == 0
            c = data['courses'].get(pr['course'],       {})
            p = data['courses'].get(pr['prerequisite'], {})
            for ci, v in enumerate([
                pr['course'],       c.get('name', ''),
                pr['type'],
                pr['prerequisite'], p.get('name', ''),
                pr.get('source', ''),
            ], 1):
                self._cell(ws, ri, ci, v, alt)

    def _sheet_credit_dist(self, wb, data):
        ws = wb.create_sheet("Phân bổ tín chỉ")
        for ci, (h, w) in enumerate(zip(
                ["Hợp phần","Số tín chỉ","Tỷ lệ (%)"], [25, 15, 15]), 1):
            self._hdr(ws, 1, ci, h, w)

        for ri, (cat, info) in enumerate(data['credit_distribution'].items(), 2):
            alt = ri % 2 == 0
            self._cell(ws, ri, 1, cat,                       alt)
            self._cell(ws, ri, 2, info['credits'],            alt)
            self._cell(ws, ri, 3, f"{info['percentage']:.1f}%", alt)

    # ── export ────────────────────────────────────────────────────────────────

    def export(self, data: dict, output_path: str):
        print(f"\n📊 Xuất Excel: {output_path}")
        wb = Workbook()
        wb.remove(wb.active)

        self._sheet_overview(wb, data)
        self._sheet_courses(wb, data)
        self._sheet_plo(wb, data)
        self._sheet_semesters(wb, data)
        self._sheet_prerequisites(wb, data)
        if data['credit_distribution']:
            self._sheet_credit_dist(wb, data)

        wb.save(output_path)
        print(f"   ✓ Đã lưu: {output_path}")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  PDF REPORTER                                                               ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

class PDFReporter:
    """Xuất báo cáo tóm tắt curriculum ra file .pdf."""

    NAVY  = colors.HexColor("#1F4E79")
    BLUE  = colors.HexColor("#2E75B6")
    LGRAY = colors.HexColor("#DEEAF1")

    def _styles(self):
        ss = getSampleStyleSheet()
        return {
            'title':   ParagraphStyle('Title2',  parent=ss['Title'],
                                      textColor=self.NAVY, fontSize=18, spaceAfter=6),
            'h1':      ParagraphStyle('H1',      parent=ss['Heading1'],
                                      textColor=self.NAVY, fontSize=14,
                                      spaceBefore=12, spaceAfter=4),
            'h2':      ParagraphStyle('H2',      parent=ss['Heading2'],
                                      textColor=self.BLUE, fontSize=11,
                                      spaceBefore=8, spaceAfter=3),
            'body':    ParagraphStyle('Body',    parent=ss['Normal'],
                                      fontSize=9, leading=13),
            'caption': ParagraphStyle('Caption', parent=ss['Normal'],
                                      fontSize=8, textColor=colors.grey,
                                      alignment=TA_CENTER),
        }

    def _tbl_style(self):
        return TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  self.NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, self.LGRAY]),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.grey),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])

    def export(self, data: dict, output_path: str):
        print(f"\n📄 Xuất PDF: {output_path}")
        doc   = SimpleDocTemplate(output_path, pagesize=A4,
                                  leftMargin=2*cm, rightMargin=2*cm,
                                  topMargin=2*cm,  bottomMargin=2*cm)
        st    = self._styles()
        story = []
        W     = doc.width
        info  = data['program_info']

        # ── Bìa ──────────────────────────────────────────────────────────────
        story += [
            Spacer(1, 2*cm),
            Paragraph("BÁO CÁO CHƯƠNG TRÌNH ĐÀO TẠO", st['title']),
            HRFlowable(width="100%", thickness=2, color=self.NAVY),
            Spacer(1, 0.3*cm),
            Paragraph(info.get('program_name', ''), st['h1']),
            Paragraph(
                f"Mã ngành: <b>{info.get('program_code','')}</b> &nbsp;|&nbsp; "
                f"Tổng tín chỉ: <b>{info.get('total_credits','')}</b> &nbsp;|&nbsp; "
                f"Năm: <b>{info.get('year','')}</b>",
                st['body']),
            Spacer(1, 1*cm),
        ]

        # ── 1. Thống kê ───────────────────────────────────────────────────────
        story.append(Paragraph("1. Thống kê tổng quan", st['h1']))
        t = Table([
            ["Chỉ số", "Số lượng"],
            ["Chuẩn đầu ra (PLO)",       str(len(data['learning_outcomes']))],
            ["Chỉ số thực hiện (PI)",    str(len(data['program_indicators']))],
            ["Học phần",                 str(len(data['courses']))],
            ["Học kỳ",                   str(len(data['semesters']))],
            ["Quan hệ tiên quyết",       str(len(data['prerequisites']))],
            ["Vị trí việc làm",          str(len(data['job_positions']))],
        ], colWidths=[W*0.65, W*0.35])
        t.setStyle(self._tbl_style())
        story += [t, Spacer(1, 0.5*cm)]

        # ── 2. PLO ───────────────────────────────────────────────────────────
        story.append(Paragraph("2. Chuẩn đầu ra chương trình (PLO)", st['h1']))
        rows = [["Mã", "Mô tả", "Loại"]]
        for plo in data['learning_outcomes']:
            desc = plo['description']
            rows.append([plo['code'],
                         Paragraph((desc[:90]+"…") if len(desc)>90 else desc, st['body']),
                         plo['type']])
        t = Table(rows, colWidths=[W*0.08, W*0.65, W*0.27])
        t.setStyle(self._tbl_style())
        story += [t, Spacer(1, 0.5*cm)]

        # ── 3. Môn học ───────────────────────────────────────────────────────
        story += [PageBreak(), Paragraph("3. Danh sách học phần (30 môn đầu)", st['h1'])]
        rows = [["Mã HP", "Tên học phần", "TC", "Loại", "Tiên quyết"]]
        for cid, c in list(data['courses'].items())[:30]:
            rows.append([cid,
                         Paragraph(c['name'], st['body']),
                         str(c['credits']),
                         "TC" if c['is_elective'] else "BB",
                         c.get('prerequisite', '') or "—"])
        t = Table(rows, colWidths=[W*0.1, W*0.48, W*0.06, W*0.07, W*0.12])
        t.setStyle(self._tbl_style())
        story += [t, Spacer(1, 0.3*cm),
                  Paragraph(f"(Tổng cộng {len(data['courses'])} học phần trong CTĐT)",
                             st['caption'])]

        # ── 4. Học kỳ ────────────────────────────────────────────────────────
        story += [PageBreak(), Paragraph("4. Phân bổ học phần theo học kỳ", st['h1'])]
        for sem in sorted(data['semesters'].values(), key=lambda s: s['number']):
            story.append(Paragraph(
                f"{sem['name']} — Năm {sem['year']} — Tổng: {sem['total_credits']} TC",
                st['h2']))
            rows = [["Mã HP", "Tên học phần", "Tín chỉ"]]
            for cid in sem['courses']:
                c = data['courses'].get(cid, {})
                rows.append([cid,
                              Paragraph(c.get('name', ''), st['body']),
                              str(c.get('credits', ''))])
            t = Table(rows, colWidths=[W*0.12, W*0.72, W*0.1])
            t.setStyle(self._tbl_style())
            story += [t, Spacer(1, 0.3*cm)]

        # ── 5. Vị trí việc làm ───────────────────────────────────────────────
        if data['job_positions']:
            story += [PageBreak(),
                      Paragraph("5. Vị trí việc làm sau tốt nghiệp", st['h1'])]
            rows = [["Vị trí", "Nhóm", "Mô tả"]]
            for j in data['job_positions']:
                rows.append([j['name'], j['category'],
                              Paragraph(j.get('description', ''), st['body'])])
            t = Table(rows, colWidths=[W*0.28, W*0.25, W*0.47])
            t.setStyle(self._tbl_style())
            story += [t, Spacer(1, 0.5*cm)]

        # ── 6. Phân bổ tín chỉ ───────────────────────────────────────────────
        if data['credit_distribution']:
            story.append(Paragraph("6. Phân bổ tín chỉ theo hợp phần", st['h1']))
            rows = [["Hợp phần", "Tín chỉ", "Tỷ lệ (%)"]]
            for cat, info in data['credit_distribution'].items():
                rows.append([cat, str(info['credits']), f"{info['percentage']:.1f}%"])
            t = Table(rows, colWidths=[W*0.5, W*0.25, W*0.25])
            t.setStyle(self._tbl_style())
            story.append(t)

        doc.build(story)
        print(f"   ✓ Đã lưu: {output_path}")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  MAIN                                                                       ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

def main():
    print("=" * 65)
    print("  KNOWLEDGE GRAPH — HCMUE CNTT")
    print("  (GPT extraction + Export Excel/PDF + Neo4j)")
    print("=" * 65)

    if not NEO4J_PASSWORD:
        print("⚠️  NEO4J_PASSWORD chưa được set  →  export NEO4J_PASSWORD=...")

    # ── 0. GPT extractor ─────────────────────────────────────────────────────
    print("\n[0] Khởi tạo GPT extractor...")
    gpt = GPTExtractor(api_key=GEMINI_API_KEY)

    # ── 1. Trích xuất PDF ────────────────────────────────────────────────────
    print("\n[1] Trích xuất dữ liệu từ PDF...")
    extractor = CurriculumExtractor(PDF_FILE_PATH, gpt_extractor=gpt)
    data = extractor.run()
    extractor.save_to_json('curriculum_extracted.json')
    extractor.print_summary()

    # ── 2. Export Excel ──────────────────────────────────────────────────────
    print("\n[2] Xuất báo cáo Excel...")
    ExcelReporter().export(data, 'curriculum_report.xlsx')

    # ── 3. Export PDF ────────────────────────────────────────────────────────
    print("\n[3] Xuất báo cáo PDF...")
    PDFReporter().export(data, 'curriculum_report.pdf')

    # ── 4. Import Neo4j ──────────────────────────────────────────────────────
    print("\n[4] Import dữ liệu vào Neo4j...")
    importer = Neo4jImporter(NEO4J_URL, NEO4J_USER, NEO4J_PASSWORD)
    try:
        importer.import_curriculum_data(data)
    finally:
        importer.close()

    print("\n" + "=" * 65)
    print("  HOÀN THÀNH!")
    print("  • curriculum_extracted.json")
    print("  • curriculum_report.xlsx")
    print("  • curriculum_report.pdf")
    print("=" * 65)


if __name__ == "__main__":
    main()